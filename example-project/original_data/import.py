import re
import pprint
from openpyxl import load_workbook

max_rows = 100
filename = 'orders.xlsx'

pp = pprint.PrettyPrinter(indent=4, width=100)

# Load our workbook and sheet
wb = load_workbook(filename=filename)
sheet = wb['Orders']

# Make sure it's the format we expect
if (sheet['A1'].value != 'Customer' or
    sheet['B1'].value != 'Parts' or
    sheet['C1'].value != 'Quantity'):
    print("Invalid format, exiting.")
    exit(1)

# Find all of the unique parts
# NOTE: Rows and Columns start at 1 in openpyxl
parts = []
for row in range(2, max_rows):
    part_name = sheet.cell(column=2, row=row).value
    if part_name and part_name not in parts:
        parts.append(part_name)
print("Unique parts detected:")
print(parts)

# Find all of the unique customers
# NOTE: customer names are spaced 5 rows apart
customers = {} # Using a dictionary helps prevent repeats
for row in range(2, max_rows, 5):
    customer_name = sheet.cell(column=1, row=row).value
    if customer_name and customer_name not in customers:
        street_address = sheet.cell(column=1, row=row+1).value
        city_state_zip = sheet.cell(column=1, row=row+2).value
        phone = sheet.cell(column=1, row=row+3).value
        # source: https://stackoverflow.com/questions/35784962/regex-for-capturing-city-state-zip-from-an-address-string?rq=1
        m = re.search('([^,]+), ([A-Z]{2}) (\d{5})', city_state_zip)
        city = m.group(1)
        state = m.group(2)
        postzip = m.group(3) # NOTE: zip is a reserved word
        customers[customer_name] = {
            'name': customer_name,
            'street_address': street_address,
            'city': city,
            'state': state,
            'zip': postzip,
            'phone': phone,
        }
print("Unique customers detected:")
pp.pprint(customers)
